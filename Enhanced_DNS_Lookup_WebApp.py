# Enhanced_DNS_Lookup_WebApp.py
# Flask-based DNS Lookup Web Application
# Upload CSV, process DNS lookups, download ZIP of results

from flask import Flask, request, render_template_string, send_file, send_from_directory, redirect
from markupsafe import Markup
import os
import zipfile
import io
import pandas as pd
import dns.resolver
import re
import whois
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
from concurrent.futures import ThreadPoolExecutor
from PIL import Image, ImageOps
from fpdf import FPDF
import uuid

# --- Fix: matplotlib backend for Flask (headless) ---
import matplotlib
matplotlib.use('Agg')  # Use non-GUI backend for chart generation

app = Flask(__name__)

# --- Enhanced Homepage ---
UPLOAD_FORM = '''
<!doctype html>
<html lang="en">
<head>
  <title>Kenvue DNS Lookup Dashboard</title>
  <style>
    body { font-family: 'Segoe UI', Arial, sans-serif; background: linear-gradient(135deg, #e3f0ff 0%, #e6f7ec 100%); min-height: 100vh; margin:0; }
    .main-header { display:flex; flex-direction:column; align-items:center; justify-content:center; background: linear-gradient(90deg, #008A4B 0%, #6FCF97 100%); color: #fff; gap: 0px; padding: 0px 0 0 0; font-size:2em; font-weight:700; letter-spacing:1px; border-radius:0 0 24px 24px; box-shadow:0 2px 16px rgba(0,0,0,0.10); }
    .kenvue-logo-top { height:156px; margin-bottom:0px; margin-top:0px; border-radius:0px; clip-path: inset(10% 0 10% 0); background:transparent; box-shadow:none; display:block; }
    .main-header-text { width:100%; text-align:center; font-size:2em; font-weight:700; margin-top:0px; margin-bottom:3px; line-height: 1.2; letter-spacing:1px; padding:20px 0 14px 0; }
    .subtitle { text-align:center; color:#008A4B; font-size:1.35em; margin-top:40px; margin-bottom:64px; font-weight:500; }
    .container { max-width: 540px; margin: 0 auto 0 auto; background: #fff; padding: 40px 36px 32px 36px; border-radius: 18px; box-shadow: 0 4px 24px rgba(0,0,0,0.10); display: flex; flex-direction: column; align-items: center; }
    .info-box { background:#e6f7ec; color:#222; border-radius:10px; padding:18px 24px; margin-bottom:48px; font-size:1.08em; box-shadow:0 2px 8px rgba(0,0,0,0.07); text-align: center; }
    .upload-label, .sample-link { text-align: center; display: inline; width: auto; }
    .upload-row { width:100%; text-align:center; margin-bottom:18px; }
    form, input[type=file], button { text-align: center; width: 100%; }
    form { width: 100%; display: flex; flex-direction: column; align-items: center; }
    input[type=file] { margin-bottom: 28px; margin-top: 18px; font-size:1em; display:block; text-align:center; margin-left:30%; margin-right:0; width: 60%; }
    button { background: linear-gradient(90deg, #008A4B 0%, #6FCF97 100%); color: #fff; border: none; padding: 14px 32px; border-radius: 8px; font-size: 1.15em; font-weight:600; cursor: pointer; box-shadow:0 2px 8px rgba(0,0,0,0.08); transition: background 0.2s; }
    button:hover { background: #005A2C; }
    .footer { text-align:center; color:#888; font-size:1em; margin-top:48px; }
    .spinner { display: none; margin: 18px auto 0 auto; border: 6px solid #e6f7ec; border-top: 6px solid #008A4B; border-radius: 50%; width: 44px; height: 44px; animation: spin 1s linear infinite; }
    @keyframes spin { 100% { transform: rotate(360deg); } }
    .success-message { display: none; color: #008A4B; font-size: 1.15em; font-weight: 600; margin-top: 22px; margin-bottom: 0px; text-align: center; }
    .domain-preview { display: none; background: #f6fff9; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.07); padding: 18px 18px; margin-bottom: 18px; font-size: 1.08em; text-align: center; }
  </style>
</head>
<body>
  <div class="main-header">
    <img src="https://upload.wikimedia.org/wikipedia/commons/thumb/3/35/Kenvue_Logo_Black_RGB.svg/2560px-Kenvue_Logo_Black_RGB.svg.png" class="kenvue-logo-top" alt="Kenvue Logo"/>
    <div class="main-header-text">Kenvue DNS Lookup Dashboard</div>
  </div>
  <div class="subtitle">Professional DNS, DMARC, SPF, MX & WHOIS Analysis for Your Domains</div>
  <div class="container">
    <div class="info-box">Upload your domain list (CSV format). Get instant dashboard, PDF, and Excel results with actionable insights. All data is processed securely and privately.</div>
    <form id="uploadForm" method="post" action="/process" enctype="multipart/form-data" onsubmit="return handleFormSubmit(event)">
      <div class="upload-row">
        <label class="upload-label">Upload Domains CSV:</label>
        <a class="sample-link" href="/download-sample" target="_blank">Download Sample File</a>
      </div>
      <input type="file" name="domains_csv" id="domains_csv" accept=".csv" required onchange="showDomainPreview(event)"><br>
      <div class="domain-preview" id="domainPreview" style="display:none;"></div>
      <button type="submit" id="submitBtn">Generate Report</button>
      <div class="spinner" id="spinner"></div>
      <div class="success-message" id="successMsg">Report generated successfully! Redirecting...</div>
    </form>
  </div>
  <div class="footer">&copy; 2025 Kenvue | Powered by Python & Flask</div>
  <script>
    function showDomainPreview(e) {
      const file = e.target.files[0];
      if (!file) { document.getElementById('domainPreview').style.display = 'none'; return; }
      const reader = new FileReader();
      reader.onload = function(evt) {
        const lines = evt.target.result.split(/\r?\n/).filter(l => l.trim());
        let preview = '<b>Domain Preview:</b><br><ul style="list-style:none;padding:0;margin:0;">';
        let count = 0;
        for (let i = 0; i < lines.length && count < 5; i++) {
          let domain = lines[i].replace(/\r|\n/g, '').trim();
          if (domain && (domain.toLowerCase().includes('domain') === false || lines.length === 1)) {
            preview += `<li>${domain}</li>`;
            count++;
          }
        }
        preview += '</ul>';
        document.getElementById('domainPreview').innerHTML = preview;
        document.getElementById('domainPreview').style.display = 'block';
      };
      reader.readAsText(file);
    }
    function handleFormSubmit(e) {
      e.preventDefault();
      document.getElementById('submitBtn').disabled = true;
      document.getElementById('spinner').style.display = 'block';
      document.getElementById('successMsg').style.display = 'none';
      var formData = new FormData(document.getElementById('uploadForm'));
      fetch('/process', { method: 'POST', body: formData })
        .then(response => {
          if (response.redirected) {
            document.getElementById('spinner').style.display = 'none';
            document.getElementById('successMsg').style.display = 'block';
            setTimeout(() => { window.location.href = response.url; }, 1200);
          } else {
            document.getElementById('spinner').style.display = 'none';
            document.getElementById('submitBtn').disabled = false;
            alert('Error generating report.');
          }
        })
        .catch(() => {
          document.getElementById('spinner').style.display = 'none';
          document.getElementById('submitBtn').disabled = false;
          alert('Error generating report.');
        });
      return false;
    }
    document.getElementById('domains_csv').addEventListener('change', showDomainPreview);
  </script>
</body>
</html>
'''

@app.route('/', methods=['GET'])
def index():
    return render_template_string(UPLOAD_FORM)

# --- Enhanced Results Page ---
@app.route('/results/<job_id>')
def results(job_id):
    temp_dir = os.path.join('webapp_results', job_id)
    dashboard_dir = os.path.join(temp_dir, 'Dashboard')
    html_files = [f for f in os.listdir(dashboard_dir) if f.endswith('.html')]
    pdf_files = [f for f in os.listdir(dashboard_dir) if f.endswith('.pdf')]
    excel_files = [f for f in os.listdir(temp_dir) if f.endswith('.xlsx')]
    dashboard_html = ''
    if html_files:
        with open(os.path.join(dashboard_dir, html_files[0]), 'r', encoding='utf-8') as f:
            dashboard_html = f.read().replace('{job_id}', job_id)
    # Results page template
    return render_template_string('''
    <!doctype html>
    <html lang="en">
    <head>
      <title>DNS Lookup Results</title>
      <style>
        body { font-family: 'Segoe UI', Arial, sans-serif; background: #f6fff9; margin:0; }
        .main-header { text-align:center; background: linear-gradient(90deg, #008A4B 0%, #6FCF97 100%); color: #fff; padding: 36px 0 18px 0; font-size:2.2em; font-weight:700; letter-spacing:1px; border-radius:0 0 18px 18px; box-shadow:0 2px 12px rgba(0,0,0,0.10); }
        .container { max-width: 1200px; margin: 40px auto 0 auto; background: #fff; padding: 36px 36px 32px 36px; border-radius: 18px; box-shadow: 0 4px 24px rgba(0,0,0,0.10); }
        .file-links { margin: 32px 0 0 0; text-align:center; }
        .file-btn { display:inline-block; margin:0 18px 18px 0; padding:12px 28px; background:#008A4B; color:#fff; border-radius:7px; text-decoration:none; font-size:1.1em; font-weight:600; box-shadow:0 2px 8px rgba(0,0,0,0.08); transition: background 0.2s; }
        .file-btn:hover { background:#005A2C; }
        .dashboard-embed { border:2px solid #008A4B; border-radius:12px; box-shadow:0 2px 12px rgba(0,0,0,0.10); margin-bottom:32px; }
      </style>
    </head>
    <body>
      <div class="main-header">DNS Lookup Results</div>
      <div class="container">
        <div class="dashboard-embed">{{ dashboard_html|safe }}</div>
        <div class="file-links">
          {% if pdf_files %}
            <a class="file-btn" href="/download/{{ job_id }}/pdf" target="_blank">Open PDF</a>
            <a class="file-btn" href="/download/{{ job_id }}/pdf?dl=1">Download PDF</a>
          {% endif %}
          {% if excel_files %}
            <a class="file-btn" href="/download/{{ job_id }}/excel" target="_blank">Open Excel</a>
            <a class="file-btn" href="/download/{{ job_id }}/excel?dl=1">Download Excel</a>
          {% endif %}
        </div>
      </div>
    </body>
    </html>
    ''', dashboard_html=Markup(dashboard_html), job_id=job_id, pdf_files=pdf_files, excel_files=excel_files)

@app.route('/results/<job_id>/image/<filename>')
def serve_dashboard_image(job_id, filename):
    temp_dir = os.path.join('webapp_results', job_id, 'Images')
    path = os.path.join(temp_dir, filename)
    if os.path.exists(path):
        return send_file(path)
    return "Image not found", 404

@app.route('/download/<job_id>/<filetype>')
def download_file(job_id, filetype):
    temp_dir = os.path.join('webapp_results', job_id)
    dashboard_dir = os.path.join(temp_dir, 'Dashboard')
    if filetype == 'pdf':
        files = [f for f in os.listdir(dashboard_dir) if f.endswith('.pdf')]
        if files:
            path = os.path.join(dashboard_dir, files[0])
            if 'dl' in request.args:
                return send_file(path, as_attachment=True)
            else:
                return send_file(path)
    elif filetype == 'excel':
        files = [f for f in os.listdir(temp_dir) if f.endswith('.xlsx')]
        if files:
            path = os.path.join(temp_dir, files[0])
            if 'dl' in request.args:
                return send_file(path, as_attachment=True)
            else:
                return send_file(path)
    return "File not found", 404

@app.route('/download-sample')
def download_sample():
    sample_path = os.path.join(os.path.dirname(__file__), 'SampleDomainList.csv')
    return send_file(sample_path, as_attachment=True)

@app.route('/process', methods=['POST'])
def process():
    file = request.files.get('domains_csv')
    if not file:
        return "No file uploaded", 400
    job_id = str(uuid.uuid4())
    temp_dir = os.path.join('webapp_results', job_id)
    os.makedirs(temp_dir, exist_ok=True)
    temp_csv = os.path.join(temp_dir, 'uploaded_domains.csv')
    file.save(temp_csv)
    run_dns_lookup(temp_csv, temp_dir)
    return redirect(f'/results/{job_id}')

# --- DNS Lookup Logic as Function ---
def run_dns_lookup(input_csv_path, output_dir):
    # Setup output folders
    images_dir = os.path.join(output_dir, "Images")
    dashboard_dir = os.path.join(output_dir, "Dashboard")
    logs_dir = os.path.join(output_dir, "Logs")
    for d in [images_dir, dashboard_dir, logs_dir]:
        os.makedirs(d, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    human_timestamp = datetime.now().strftime("%B %d, %Y at %I:%M %p IST")
    log_file = os.path.join(logs_dir, f"DNS_Script_Logs_{timestamp}.txt")
    html_file = os.path.join(dashboard_dir, f"DNS_Lookup_Summary_Dashboard_{timestamp}.html")
    final_output_file = os.path.join(output_dir, f"DNS_Lookup_Results_{timestamp}.xlsx")
    pdf_file = os.path.join(dashboard_dir, f"DNS_Lookup_Report_{timestamp}.pdf")
    logging.basicConfig(filename=log_file, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    # Load domains
    df = pd.read_csv(input_csv_path)
    # Workbook setup
    wb = Workbook()
    ws_dmarc = wb.active; ws_dmarc.title = "DMARC"
    ws_spf = wb.create_sheet("SPF")
    ws_mx = wb.create_sheet("MX")
    ws_whois = wb.create_sheet("WHOIS")
    ws_summary = wb.create_sheet("Summary")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    light_red_fill = PatternFill(start_color='FF4433', end_color='FF4433', fill_type='solid')
    light_green_fill = PatternFill(start_color='4CBB17', end_color='4CBB17', fill_type='solid')
    light_blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
    orange_fill = PatternFill(start_color='FBCEB1', end_color='FBCEB1', fill_type='solid')
    light_yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    bold_font = Font(bold=True)
    center_align = Alignment(wrap_text=True, vertical='center')
    ws_dmarc.append(["Domain", "Primary_Domain_Policy", "Secondary_Domain_Policy", "DMARC_Record"])
    ws_spf.append(["Domain", "SPF_Record"])
    ws_mx.append(["Domain", "MX_Record"])
    ws_whois.append(["Domain", "NameServers", "Registrar", "RegisteredOn", "ExpiresOn", "UpdatedOn"])
    for sheet in [ws_dmarc, ws_spf, ws_mx, ws_whois, ws_summary]:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_align
        sheet.auto_filter.ref = sheet.dimensions
    spf_chart_data = {"No SPF Record": 0, "Explicit Hard Fail": 0, "JNJ Agari SPF": 0, "Third Party SPF": 0}
    mx_chart_data = {"No MX Record": 0, "Kenvue MX": 0, "JNJ MX": 0, "Third Party MX": 0}
    dmarc_ownership = {"No DMARC Record": 0, "Non-Migrated JNJ DMARC": 0, "Migrated Kenvue DMARC": 0}
    dmarc_policy = {"No DMARC Record": 0, "Reject DMARC Policy": 0, "Quarantine DMARC Policy": 0, "No DMARC Policy": 0}
    whois_chart_data = {"No Name Servers Found": 0, "Kenvue Owned Domains": 0, "Non-Kenvue Domain": 0}
    def get_dns_record(domain, record_type, nameserver='1.1.1.1'):
        resolver = dns.resolver.Resolver()
        resolver.nameservers = [nameserver]
        resolver.timeout = 10
        resolver.lifetime = 20
        try:
            answers = resolver.resolve(domain, record_type)
            return ', '.join(answer.to_text() for answer in answers)
        except:
            return f'No {record_type} record found'
    def get_spf_record(domain, nameserver='1.1.1.1'):
        resolver = dns.resolver.Resolver()
        resolver.nameservers = [nameserver]
        resolver.timeout = 10
        resolver.lifetime = 20
        try:
            answers = resolver.resolve(domain, 'TXT')
            for rdata in answers:
                for txt_string in rdata.strings:
                    decoded = txt_string.decode('utf-8')
                    if decoded.startswith('v=spf1'):
                        return decoded
            return 'No SPF record found'
        except:
            return 'No SPF record found'
    def extract_policy(dmarc_record, policy_type):
        match = re.search(policy_type + '=([^;]+)', dmarc_record)
        return match.group(1) if match else f'No {policy_type} policy found'
    def format_date(date_obj):
        if isinstance(date_obj, list):
            date_obj = date_obj[0]
        if isinstance(date_obj, datetime):
            return date_obj.strftime("%Y-%m-%d")
        return ""
    def normalize_nameservers(ns_list):
        return [re.sub(r'\s+', '', ns.strip().lower()) for ns in ns_list if ns.strip()]
    def process_domain(domain):
        logging.info(f"Processing domain: {domain}")
        dmarc_record = get_dns_record(f"_dmarc.{domain}", "TXT")
        p_policy = extract_policy(dmarc_record, 'p')
        sp_policy = extract_policy(dmarc_record, 'sp')
        if 'No TXT record found' in dmarc_record:
            dmarc_ownership["No DMARC Record"] += 1
            dmarc_policy["No DMARC Record"] += 1
        else:
            if "jnj@rua.dmp.cisco.com" in dmarc_record or "jnj@ruf.dmp.cisco.com" in dmarc_record:
                dmarc_ownership["Non-Migrated JNJ DMARC"] += 1
            elif "93881cb5@inbox.ondmarc.com" in dmarc_record:
                dmarc_ownership["Migrated Kenvue DMARC"] += 1
            if 'p=reject' in dmarc_record:
                dmarc_policy["Reject DMARC Policy"] += 1
            elif 'p=quarantine' in dmarc_record:
                dmarc_policy["Quarantine DMARC Policy"] += 1
            elif 'p=none' in dmarc_record:
                dmarc_policy["No DMARC Policy"] += 1
        ws_dmarc.append([domain, p_policy, sp_policy, dmarc_record])
        row_dmarc = ws_dmarc.max_row
        for cell in ws_dmarc[row_dmarc]:
            cell.border = thin_border
            cell.alignment = center_align
            if dmarc_record == 'No TXT record found':
                cell.fill = light_red_fill
            elif p_policy == 'quarantine' and sp_policy == 'quarantine':
                cell.fill = light_green_fill
            elif p_policy == 'reject' and sp_policy == 'reject':
                cell.fill = light_blue_fill
            elif p_policy == 'none' and sp_policy == 'none':
                cell.fill = orange_fill
            if "rua=mailto:jnj@rua.dmp.cisco.com" in dmarc_record.lower() or "ruf=mailto:jnj@ruf.dmp.cisco.com" in dmarc_record.lower():
                cell.fill = light_yellow_fill
        spf_record = get_spf_record(domain)
        ws_spf.append([domain, spf_record])
        row_spf = ws_spf.max_row
        for cell in ws_spf[row_spf]:
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = light_green_fill if 'No' not in spf_record else light_red_fill
        if 'No SPF record found' in spf_record:
            spf_chart_data["No SPF Record"] += 1
        elif spf_record.strip() == 'v=spf1 -all':
            spf_chart_data["Explicit Hard Fail"] += 1
        elif "ce.spf-protect.dmp.cisco.com" in spf_record or "d.espf.dmp.cisco.com" in spf_record:
            spf_chart_data["JNJ Agari SPF"] += 1
        else:
            spf_chart_data["Third Party SPF"] += 1
        mx_record = get_dns_record(domain, 'MX')
        ws_mx.append([domain, mx_record])
        row_mx = ws_mx.max_row
        for cell in ws_mx[row_mx]:
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = light_green_fill if 'No' not in mx_record else light_red_fill
        if 'No MX record found' in mx_record:
            mx_chart_data["No MX Record"] += 1
        elif "kenvue-com.mail.protection.outlook.com" in mx_record:
            mx_chart_data["Kenvue MX"] += 1
        elif "mx1.jnj-sd.iphmx.com" in mx_record or "mx2.jnj-sd.iphmx.com" in mx_record:
            mx_chart_data["JNJ MX"] += 1
        else:
            mx_chart_data["Third Party MX"] += 1
        try:
            w = whois.whois(domain)
            ns_list = w.name_servers if isinstance(w.name_servers, list) else [w.name_servers] if isinstance(w.name_servers, str) else []
            ns_display = "\n".join(ns_list)
            normalized_ns = normalize_nameservers(ns_list)
            if not normalized_ns:
                whois_chart_data["No Name Servers Found"] += 1
            elif any("kenvuedns" in ns for ns in normalized_ns):
                whois_chart_data["Kenvue Owned Domains"] += 1
            else:
                whois_chart_data["Non-Kenvue Domain"] += 1
            ws_whois.append([domain, ns_display, w.registrar, format_date(w.creation_date), format_date(w.expiration_date), format_date(w.updated_date)])
            row_whois = ws_whois.max_row
            fill = light_green_fill if any("kenvuedns" in ns for ns in normalized_ns) else light_red_fill if not normalized_ns else light_yellow_fill
            for cell in ws_whois[row_whois]:
                cell.border = thin_border
                cell.alignment = center_align
                cell.fill = fill
        except Exception as e:
            whois_chart_data["No Name Servers Found"] += 1
            ws_whois.append([domain, f"Error: {e}", "", "", "", ""])
            row_whois = ws_whois.max_row
            for cell in ws_whois[row_whois]:
                cell.border = thin_border
                cell.alignment = center_align
                cell.fill = light_red_fill
    with ThreadPoolExecutor(max_workers=10) as executor:
        for _ in executor.map(process_domain, df["Domain"]):
            pass
    for ws in [ws_dmarc, ws_spf, ws_mx, ws_whois]:
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    def create_and_embed_chart(data, title, filename, position):
        labels = list(data.keys())
        sizes = list(data.values())
        plt.figure(figsize=(8, 6))
        plt.pie(sizes, labels=[f"{l} ({v} / {sum(sizes)} - {v/sum(sizes)*100:.1f}%)" for l, v in data.items()], autopct='', startangle=140)
        plt.title(title, fontsize=14, pad=20)
        chart_path = os.path.join(images_dir, filename)
        plt.savefig(chart_path, bbox_inches='tight')
        plt.close()
        img = Image.open(chart_path)
        bordered = ImageOps.expand(img, border=6, fill='black')
        bordered.save(chart_path)
        xl_img = XLImage(chart_path)
        xl_img.width = 500
        xl_img.height = 400
        ws_summary.add_image(xl_img, position)
    create_and_embed_chart(spf_chart_data, "SPF Summary", "spf_chart.png", "A1")
    create_and_embed_chart(mx_chart_data, "MX Summary", "mx_chart.png", "J1")
    create_and_embed_chart(dmarc_ownership, "DMARC Ownership", "dmarc_ownership.png", "A25")
    create_and_embed_chart(dmarc_policy, "DMARC Policy", "dmarc_policy.png", "J25")
    create_and_embed_chart(whois_chart_data, "WHOIS Summary", "whois_chart.png", "A49")
    wb.save(final_output_file)
    def dashboard_image_url(filename):
        return f"/results/{os.path.basename(output_dir)}/image/{filename}"
    def write_summary_table(f, title, summary_dict):
        f.write('<table class="summary-table"><tr><th>Category</th><th>Count</th></tr>')
        for k, v in summary_dict.items():
            f.write(f'<tr><td>{k}</td><td>{v}</td></tr>')
        f.write('</table>')
    def get_unique_pointers(*pointers):
        seen = set()
        unique = []
        for p in pointers:
            if p not in seen:
                unique.append(p)
                seen.add(p)
        return unique
    def get_dynamic_pointer_spf(spf_data):
        total = sum(spf_data.values())
        if spf_data["No SPF Record"] > 0:
            return f"{spf_data['No SPF Record']} domains are missing SPF records. Please review these for improved email security."
        elif spf_data["Explicit Hard Fail"] > 0:
            return f"{spf_data['Explicit Hard Fail']} domains have explicit hard fail SPF policies. Ensure this is intentional."
        elif spf_data["JNJ Agari SPF"] > 0:
            return f"{spf_data['JNJ Agari SPF']} domains use JNJ Agari SPF. Confirm these are correctly configured."
        else:
            return f"Most domains ({spf_data['Third Party SPF']}) use third-party SPF records. Review for compliance."
    def get_dynamic_pointer_mx(mx_data):
        if mx_data["No MX Record"] > 0:
            return f"{mx_data['No MX Record']} domains are missing MX records. These domains cannot receive emails."
        elif mx_data["Kenvue MX"] > 0:
            return f"{mx_data['Kenvue MX']} domains are using Kenvue MX. Ensure these are managed as expected."
        elif mx_data["JNJ MX"] > 0:
            return f"{mx_data['JNJ MX']} domains are still using JNJ MX. Review migration status."
        else:
            return f"Most domains ({mx_data['Third Party MX']}) use third-party MX records. Review for compliance."
    def get_dynamic_pointer_dmarc_ownership(dmarc_ownership):
        if dmarc_ownership["No DMARC Record"] > 0:
            return f"{dmarc_ownership['No DMARC Record']} domains lack DMARC records. Add DMARC for better protection."
        elif dmarc_ownership["Migrated Kenvue DMARC"] > 0:
            return f"{dmarc_ownership['Migrated Kenvue DMARC']} domains have migrated to Kenvue DMARC. Good progress!"
        elif dmarc_ownership["Non-Migrated JNJ DMARC"] > 0:
            return f"{dmarc_ownership['Non-Migrated JNJ DMARC']} domains still use JNJ DMARC. Review migration plan."
        else:
            return "All domains have DMARC records."
    def get_dynamic_pointer_dmarc_policy(dmarc_policy):
        if dmarc_policy["No DMARC Record"] > 0:
            return f"{dmarc_policy['No DMARC Record']} domains lack DMARC records. Add DMARC for better protection."
        elif dmarc_policy["Reject DMARC Policy"] > 0:
            return f"{dmarc_policy['Reject DMARC Policy']} domains enforce reject DMARC policy. This is recommended."
        elif dmarc_policy["Quarantine DMARC Policy"] > 0:
            return f"{dmarc_policy['Quarantine DMARC Policy']} domains enforce quarantine DMARC policy. We can consider upgrading to reject policy."
        elif dmarc_policy["No DMARC Policy"] > 0:
            return f"{dmarc_policy['No DMARC Policy']} domains have no DMARC policy. Set a policy for better protection."
        else:
            return "All domains have DMARC policies."
    def get_dynamic_pointer_whois(whois_data):
        if whois_data["No Name Servers Found"] > 0:
            return f"{whois_data['No Name Servers Found']} domains have no name servers. Review registration status and confirm with DNS Team if Kenvue actually owns these."
        elif whois_data["Kenvue Owned Domains"] > 0:
            return f"{whois_data['Kenvue Owned Domains']} domains are Kenvue owned. Good asset management!"
        elif whois_data["Non-Kenvue Domain"] > 0:
            return f"{whois_data['Non-Kenvue Domain']} domains point to Non-Kenvue DNS. Review ownership and risk."
        else:
            return "All domains have valid name servers."
    spf_pointer = get_dynamic_pointer_spf(spf_chart_data)
    mx_pointer = get_dynamic_pointer_mx(mx_chart_data)
    dmarc_ownership_pointer = get_dynamic_pointer_dmarc_ownership(dmarc_ownership)
    dmarc_policy_pointer = get_dynamic_pointer_dmarc_policy(dmarc_policy)
    whois_pointer = get_dynamic_pointer_whois(whois_chart_data)
    unique_pointers = get_unique_pointers(spf_pointer, mx_pointer, dmarc_ownership_pointer, dmarc_policy_pointer, whois_pointer)
    with open(html_file, "w") as f:
        # Use .format() only for variables, not for CSS curly braces
        html_content = '''
        <html>
        <head>
            <title>Domain Lookup Dashboard</title>
            <style>
                body {{ font-family: 'Segoe UI', 'Arial', sans-serif; text-align: center; margin: 0; padding: 0; background: linear-gradient(135deg, #e3f0ff 0%, #e6f7ec 100%); min-height: 100vh; }}
                h1 {{ text-align: center; background: linear-gradient(90deg, #008A4B 0%, #6FCF97 100%); color: white; padding: 40px 0; margin-bottom: 40px; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,0.10); font-size: 2.5em; letter-spacing: 1px; }}
                .top-summary {{ font-family: 'Segoe UI', 'Arial', sans-serif; color: #222; background: #fff; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.07); padding: 40px 60px 32px 60px; margin: 0 auto 90px auto; max-width: 1100px; font-size: 1.25em; text-align: center; }}
                .pointer-box {{ background: linear-gradient(135deg, #fffbe6 0%, #e6ffe6 100%); border-radius: 14px; box-shadow: 0 4px 18px rgba(0,0,0,0.12); padding: 40px 40px 32px 40px; margin: 0 auto 160px auto; max-width: 900px; font-size: 1.25em; color: #222; }}
                .pointer-title {{ font-size: 2em; font-weight: 700; color: #008A4B; margin-bottom: 22px; }}
                ul {{ text-align: left; margin: 0 auto; max-width: 700px; padding-left: 30px; }}
                li {{ margin-bottom: 18px; font-size: 1.15em; }}
                .summary-section {{ display: flex; justify-content: center; align-items: flex-start; gap: 320px; margin-bottom: 240px; }}
                .summary-table-container {{ display: flex; flex-direction: column; align-items: center; min-width: 400px; max-width: 500px; }}
                .summary-table-title {{ font-size: 1.35em; font-weight: 600; color: #008A4B; margin-bottom: 18px; margin-top: 0; text-align: center; }}
                .summary-table {{ border-collapse: collapse; background: #f6fff9; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,0.10); width: 100%; font-size: 1.15em; }}
                .summary-table th, .summary-table td {{ border: 1px solid #b2e5c7; padding: 14px 20px; text-align: center; }}
                .summary-table th {{ background: #008A4B; color: #fff; font-weight: bold; font-size: 1.1em; }}
                .summary-table tr:nth-child(even) {{ background: #e6f7ec; }}
                .dashboard-img-container {{ display: flex; flex-direction: column; align-items: center; min-width: 400px; max-width: 500px; }}
                .dashboard-img {{ max-width: 500px; min-width: 400px; height: 400px; border:2px solid #008A4B; margin-bottom: 10px; display: block; box-shadow: 0 2px 12px rgba(0,0,0,0.10); background: #fff; border-radius: 12px; padding: 12px; }}
                .download-btn {{ display:inline-block; padding:10px 20px; background:#008A4B; color:#fff; text-decoration:none; border-radius:7px; margin-bottom:0px; margin-top:0px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); transition: background 0.2s; font-size: 1.1em; }}
                .download-btn:hover {{ background:#005A2C; }}
                hr {{ margin:40px 0; border: none; border-top: 2px solid #b2e5c7; }}
                .tooltip {{ position: relative; display: inline-block; }}
                .tooltip .tooltiptext {{ visibility: hidden; width: 260px; background-color: #222; color: #fff; text-align: center; border-radius: 8px; padding: 10px; position: absolute; z-index: 1; bottom: 110%; left: 50%; margin-left: -130px; opacity: 0; transition: opacity 0.3s; font-size: 1em; }}
                .tooltip:hover .tooltiptext {{ visibility: visible; opacity: 1; }}
            </style>
        </head>
        <body>
            <h1>Domain Lookup Summary</h1>
            <div class="top-summary">
                <span style='display:block;margin-bottom:12px;'>This comprehensive DNS report covers <b>{total_domains}</b> domains, analyzed and visualized for Kenvue.</span>
                <span style='display:block;'>Your latest domain insights are ready! This report was generated on <b>{human_timestamp}</b></span>
            </div>
            <div class="pointer-box">
                <div class="pointer-title">Key Insights</div>
                <ul>
        '''.format(total_domains=len(df), human_timestamp=human_timestamp)
        f.write(html_content)
        for pointer in unique_pointers:
            f.write(f'<li>{pointer}</li>')
        f.write('''
                </ul>
            </div>
        ''')
        # SPF
        f.write('<div class="summary-section">')
        f.write('<div class="summary-table-container">')
        f.write('<div class="summary-table-title">SPF Summary Table</div>')
        write_summary_table(f, None, spf_chart_data)
        f.write('</div>')
        f.write('<div class="dashboard-img-container">')
        f.write(f'<span class="tooltip"><img class="dashboard-img" src="/results/{{job_id}}/image/spf_chart.png" alt="SPF Summary" title="SPF Chart: Distribution of SPF record types."/><span class="tooltiptext">SPF Chart: Shows distribution of SPF record types across domains.</span></span>')
        f.write(f'<a class="download-btn" href="/results/{{job_id}}/image/spf_chart.png" download>Download SPF Image</a>')
        f.write('</div>')
        f.write('</div><hr>')
        # MX
        f.write('<div class="summary-section">')
        f.write('<div class="summary-table-container">')
        f.write('<div class="summary-table-title">MX Summary Table</div>')
        write_summary_table(f, None, mx_chart_data)
        f.write('</div>')
        f.write('<div class="dashboard-img-container">')
        f.write(f'<span class="tooltip"><img class="dashboard-img" src="/results/{{job_id}}/image/mx_chart.png" alt="MX Summary" title="MX Chart: Distribution of MX record types."/><span class="tooltiptext">MX Chart: Shows distribution of MX record types across domains.</span></span>')
        f.write(f'<a class="download-btn" href="/results/{{job_id}}/image/mx_chart.png" download>Download MX Image</a>')
        f.write('</div>')
        f.write('</div><hr>')
        # DMARC Ownership
        f.write('<div class="summary-section">')
        f.write('<div class="summary-table-container">')
        f.write('<div class="summary-table-title">DMARC Ownership Summary Table</div>')
        write_summary_table(f, None, dmarc_ownership)
        f.write('</div>')
        f.write('<div class="dashboard-img-container">')
        f.write(f'<span class="tooltip"><img class="dashboard-img" src="/results/{{job_id}}/image/dmarc_ownership.png" alt="DMARC Ownership" title="DMARC Ownership Chart: Ownership status."/><span class="tooltiptext">DMARC Ownership Chart: Shows DMARC ownership status across domains.</span></span>')
        f.write(f'<a class="download-btn" href="/results/{{job_id}}/image/dmarc_ownership.png" download>Download DMARC Ownership Image</a>')
        f.write('</div>')
        f.write('</div><hr>')
        # DMARC Policy
        f.write('<div class="summary-section">')
        f.write('<div class="summary-table-container">')
        f.write('<div class="summary-table-title">DMARC Policy Summary Table</div>')
        write_summary_table(f, None, dmarc_policy)
        f.write('</div>')
        f.write('<div class="dashboard-img-container">')
        f.write(f'<span class="tooltip"><img class="dashboard-img" src="/results/{{job_id}}/image/dmarc_policy.png" alt="DMARC Policy" title="DMARC Policy Chart: Policy status."/><span class="tooltiptext">DMARC Policy Chart: Shows DMARC policy status across domains.</span></span>')
        f.write(f'<a class="download-btn" href="/results/{{job_id}}/image/dmarc_policy.png" download>Download DMARC Policy Image</a>')
        f.write('</div>')
        f.write('</div><hr>')
        # WHOIS
        f.write('<div class="summary-section">')
        f.write('<div class="summary-table-container">')
        f.write('<div class="summary-table-title">WHOIS Summary Table</div>')
        write_summary_table(f, None, whois_chart_data)
        f.write('</div>')
        f.write('<div class="dashboard-img-container">')
        f.write(f'<span class="tooltip"><img class="dashboard-img" src="/results/{{job_id}}/image/whois_chart.png" alt="WHOIS Summary" title="WHOIS Chart: Ownership status."/><span class="tooltiptext">WHOIS Chart: Shows domain ownership and name server status.</span></span>')
        f.write(f'<a class="download-btn" href="/results/{{job_id}}/image/whois_chart.png" download>Download WHOIS Image</a>')
        f.write('</div>')
        f.write('</div><hr>')
        f.write("</body></html>")
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    for title, chart in zip([
        "SPF SUMMARY", "MX SUMMARY", "DMARC OWNERSHIP", "DMARC POLICY", "WHOIS SUMMARY"],
        ["spf_chart.png", "mx_chart.png", "dmarc_ownership.png", "dmarc_policy.png", "whois_chart.png"]):
        chart_path = os.path.join(images_dir, chart)
        pdf.add_page()
        pdf.set_font("Arial", size=16)
        pdf.cell(200, 16, txt=title, ln=True, align='C')
        pdf.image(chart_path, x=10, y=30, w=180)
    pdf.output(pdf_file)
    # Return all output file paths for ZIP
    return [final_output_file, html_file, pdf_file,
            os.path.join(images_dir, "spf_chart.png"),
            os.path.join(images_dir, "mx_chart.png"),
            os.path.join(images_dir, "dmarc_ownership.png"),
            os.path.join(images_dir, "dmarc_policy.png"),
            os.path.join(images_dir, "whois_chart.png"),
            log_file]

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port)
